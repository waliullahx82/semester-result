"""Convert the compiled 1-1 semester spreadsheet into app result JSON."""

from __future__ import annotations

import json
import shutil
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = ROOT / "1-1 result" / "result.xlsx"
OUTPUT_PATH = ROOT / "src" / "data" / "results-1-1.generated.json"
PUBLIC_SOURCES = ROOT / "public" / "sources"
PUBLIC_COPY = PUBLIC_SOURCES / "1-1-result.xlsx"

GRADE_POINTS = {
    "A+": 4.0,
    "A": 3.75,
    "A-": 3.5,
    "B+": 3.25,
    "B": 3.0,
    "B-": 2.75,
    "C+": 2.5,
    "C": 2.25,
    "C-": 2.0,
    "F": 0.0,
}

POINT_TO_LETTER = {value: letter for letter, value in GRADE_POINTS.items()}

COURSES = [
    {
        "id": "spl-theory-1-1",
        "code": "CSE SPL Theory",
        "title": "Structured Programming Language",
        "credits": 3.0,
        "column": "SPL Theory",
    },
    {
        "id": "spl-lab-1-1",
        "code": "CSE SPL Lab",
        "title": "Structured Programming Language Lab",
        "credits": 1.5,
        "column": "SPL Lab",
    },
    {
        "id": "discrete-1-1",
        "code": "CSE Discrete",
        "title": "Discrete Mathematics",
        "credits": 3.0,
        "column": "Discrete",
    },
    {
        "id": "eee-theory-1-1",
        "code": "EEE Theory",
        "title": "Electrical / Electronic Theory",
        "credits": 3.0,
        "column": "EEE",
    },
    {
        "id": "eee-lab-1-1",
        "code": "EEE Lab",
        "title": "Electrical / Electronic Lab",
        "credits": 1.5,
        "column": "EEE Lab",
    },
    {
        "id": "math-1-1",
        "code": "MAT Math",
        "title": "Mathematics",
        "credits": 3.0,
        "column": "Math",
    },
    {
        "id": "english-1-1",
        "code": "ENG English",
        "title": "English",
        "credits": 3.0,
        "column": "Eng",
    },
    {
        "id": "english-lab-1-1",
        "code": "ENG Lab",
        "title": "English Lab",
        "credits": 1.5,
        "column": "Eng Lab",
    },
]

SOURCE_ID = "1-1-compiled-xlsx"
REGULAR_PREFIX = "202433"


def normalize_registration(value: object) -> str:
    if isinstance(value, float):
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def letter_for_point(point: float) -> str:
    if point in POINT_TO_LETTER:
        return POINT_TO_LETTER[point]
    closest = min(GRADE_POINTS.values(), key=lambda candidate: abs(candidate - point))
    return POINT_TO_LETTER[closest]


def main() -> None:
    if not SOURCE_XLSX.exists():
        raise SystemExit(f"Missing 1-1 spreadsheet: {SOURCE_XLSX}")

    workbook = load_workbook(SOURCE_XLSX, data_only=True)
    sheet = workbook.active
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    header_index = {header: index + 1 for index, header in enumerate(headers)}

    required = ["Registration No.", "Student's Name", *[course["column"] for course in COURSES]]
    missing = [name for name in required if name not in header_index]
    if missing:
        raise SystemExit(f"Unexpected 1-1 spreadsheet headers. Missing: {missing}")

    PUBLIC_SOURCES.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SOURCE_XLSX, PUBLIC_COPY)

    students: dict[str, dict] = {}
    results: list[dict] = []
    complete_regular = 0

    for row in range(2, sheet.max_row + 1):
        raw_registration = sheet.cell(row, header_index["Registration No."]).value
        if raw_registration is None:
            continue
        registration = normalize_registration(raw_registration)
        if not registration.isdigit() or len(registration) != 10:
            raise SystemExit(f"Invalid registration on row {row}: {raw_registration!r}")

        name = sheet.cell(row, header_index["Student's Name"]).value
        name_text = str(name).strip() if name else None
        is_regular = registration.startswith(REGULAR_PREFIX)
        students[registration] = {
            "registration": registration,
            "name": name_text,
            "cohort": registration[:6],
            "isRegular": is_regular,
        }

        course_points: list[float | None] = []
        for course in COURSES:
            raw_point = sheet.cell(row, header_index[course["column"]]).value
            if raw_point is None or raw_point == "":
                course_points.append(None)
                continue
            point = float(raw_point)
            course_points.append(point)
            results.append(
                {
                    "registration": registration,
                    "courseId": course["id"],
                    "letterGrade": letter_for_point(point),
                    "gradePoint": point,
                    "status": "official",
                    "sourceIds": [SOURCE_ID],
                    "score": None,
                    "maxScore": None,
                    "note": None,
                }
            )

        if is_regular and all(point is not None for point in course_points):
            complete_regular += 1

    total_credits = sum(course["credits"] for course in COURSES)
    payload = {
        "schemaVersion": 1,
        "semester": {
            "department": "Department of Computer Science & Engineering",
            "degree": "Bachelor of Science (Engineering)",
            "name": "1-1 Semester",
            "number": "1st Semester",
            "session": "2024-2025",
            "usn": "January-June",
            "regularRegistrationPrefix": REGULAR_PREFIX,
        },
        "gradePoints": GRADE_POINTS,
        "courses": [
            {
                "id": course["id"],
                "code": course["code"],
                "title": course["title"],
                "credits": course["credits"],
                "status": "official",
                "sourceIds": [SOURCE_ID],
                "semesterKey": "1-1",
            }
            for course in COURSES
        ],
        "students": sorted(students.values(), key=lambda item: item["registration"]),
        "results": results,
        "sources": [
            {
                "id": SOURCE_ID,
                "fileName": "1-1-result.xlsx",
                "url": f"/sources/{quote('1-1-result.xlsx')}",
                "kind": "xlsx",
                "status": "official",
                "courseId": COURSES[0]["id"],
                "pageCount": None,
                "rowCount": len(students),
                "printedSession": "2024-2025",
                "publishedDate": None,
                "notes": [
                    "Compiled 1-1 semester grade sheet supplied as Excel.",
                    "Course titles and credits are normalized for explorer display; letter grades are derived from grade points.",
                ],
            }
        ],
        "issues": [
            {
                "id": "1-1-compiled-source",
                "severity": "warning",
                "type": "compiled-source",
                "courseId": COURSES[0]["id"],
                "registration": None,
                "sourceIds": [SOURCE_ID],
                "message": "1-1 results come from a compiled spreadsheet rather than per-course PDFs.",
            }
        ],
        "stats": {
            "sourceCount": 1,
            "officialSourceCount": 1,
            "courseCount": len(COURSES),
            "regularStudentCount": sum(1 for student in students.values() if student["isRegular"]),
            "officialRegularStudentCount": sum(1 for student in students.values() if student["isRegular"]),
            "labCoveredRegularCount": sum(1 for student in students.values() if student["isRegular"]),
            "eligibleOverallCount": complete_regular,
            "totalCredits": total_credits,
        },
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH.relative_to(ROOT)} with {len(students)} students and {len(results)} results.")


if __name__ == "__main__":
    main()
